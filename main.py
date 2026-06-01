import asyncio  # Librería para manejar tareas asíncronas
import os  # Librería para interactuar con el sistema operativo (guardar archivos)
import re  # Librería para búsqueda de expresiones regulares
from tkinter import END, Listbox  # Librería para crear interfaces gráficas en Python

import customtkinter as ctk  # Lo mismo que tkinter pero más moderno y con mayor personalización
from bleak import BleakClient  # Librería para conectar con módules BTE desde Python
from openpyxl import (
    Workbook,
    load_workbook,
)  # Librería para trabajar con Excel desde Python

# CONFIGURACIÓN

# Nombre de la hoja de cálculo donde se guardara la información del paciente
xlfile = "data.xlsx"

# La dirección MAC del módulo bluetooth HM-10
DEVICE_ADDRESS = "50:33:8B:1C:D3:AB"

# El "canal" que usa el módulo HM-10 para mandar información en texto plano
HM10_UUID = "0000ffe1-0000-1000-8000-00805f9b34fb"

# Variables globales: Estás actuan como almacenamiento temporal para guardar la información del paciente entre la pantalla de registro y la visualización de registros
user_name = ""
user_age = ""
user_sex = ""
user_pregnant = "No"

# Personalización
ctk.set_appearance_mode("Dark")
ctk.set_default_color_theme("green")


# INFORMACIÓN Y FUNCIONES PARA MANEJO DE ARCHIVOS
def init_excel():
    """Creates the excel file with headers if it doesn't exist."""
    if not os.path.exists(xlfile):
        wb = Workbook()  # Crear un nuevo workbook de Excel
        ws = wb.active  # Seleccionar la hoja activa
        # Agregar nombres de las columnas
        ws.append(["Nombre", "Edad", "Sexo", "Embarazada", "BPM", "Riesgo"])
        wb.save(xlfile)


# Correr la inicialización de Excel inmediatamente cuando el programa inicia
init_excel()


def calculate_risk(bpm, age, sex, pregnant):
    """
    Advanced diagnostic calculation based directly on your medical tables.
    Returns: 'Normal', 'Bradicardia', 'Taquicardia', or 'Choque Cardiogénico (Alto Riesgo)'
    """
    # Convertir los latidos (BPM) y la edad a número para cálculos
    # ya que el módulo envía la información en texto plano
    bpm = int(bpm)
    age = int(age)

    # Lo primero que hace el programa es revisar si el paciente está en situación crítica,
    # sin importar su sexo o si es adulto o niño.

    # 18 años o más
    if age >= 18:
        # Si es mujer embarazada y los latidos son muy altos o bajos
        if pregnant != "No embarazada" and (bpm > 130 or bpm < 60):
            return "Choque Cardiogénico (Alto Riesgo)"
        # Cualquier otro caso (hombre y mujer no embarazada)
        elif bpm > 120 or bpm < 50:
            return "Choque Cardiogénico (Alto Riesgo)"

    # Entre 13 y 17 años
    elif 13 <= age <= 17:
        if bpm > 120 or bpm < 50:
            return "Choque Cardiogénico (Alto Riesgo)"

    # Entre 6 y 12 años
    elif 6 <= age <= 12:
        if bpm > 130 or bpm < 60:
            return "Choque Cardiogénico (Alto Riesgo)"

    # 4 o 5 años
    elif 4 <= age <= 5:
        if bpm > 140 or bpm < 70:
            return "Choque Cardiogénico (Alto Riesgo)"

    # Entre 1 y 3 años
    elif 1 <= age <= 3:
        if bpm > 150 or bpm < 70:
            return "Choque Cardiogénico (Alto Riesgo)"

    # Menor de 1 año
    elif age < 1:
        if bpm > 160 or bpm < 70:
            return "Choque Cardiogénico (Alto Riesgo)"

    # Si el paciente no estuvo en peligro de choque en la parte anterior, se checa si es
    # una mujer embarazada, ya que sus rangos normales cambian según el trimestre.
    if sex == "Femenino" and pregnant != "No embarazada":
        # Reglas para el Primer Trimestre
        if pregnant == "Primer trimestre":
            if bpm < 60:
                return "Bradicardia"
            elif 70 <= bpm <= 100:
                return "Normal"
            elif bpm > 100:
                return "Taquicardia"

        # Reglas para el Segundo Trimestre
        elif pregnant == "Segundo trimestre":
            if bpm < 60:
                return "Bradicardia"
            elif 75 <= bpm <= 105:
                return "Normal"
            elif bpm > 105:
                return "Taquicardia"

        # Reglas para el Tercer Trimestre
        elif pregnant == "Tercer trimestre":
            if bpm < 60:
                return "Bradicardia"
            elif 80 <= bpm <= 110:
                return "Normal"
            elif bpm > 110:
                return "Taquicardia"

        # Reglas para Embarazo Avanzado
        elif pregnant == "Embarazo avanzado":
            if bpm < 60:
                return "Bradicardia"
            elif 85 <= bpm <= 115:
                return "Normal"
            elif bpm > 115:
                return "Taquicardia"

        # Colchón de seguridad: Si los latidos caen en un pequeño "espacio" de la tabla
        # (por ejemplo, 65 BPM en el primer trimestre), el sistema asume que es "Normal".
        return "Normal"

    # Si el paciente no está en riesgo de choque y no es una mujer embarazada,
    # entra a esta última sección para evaluar según su grupo de edad y sexo.

    # Niños de 1 a 3 años (Ambos sexos)
    if 1 <= age <= 3:
        if bpm < 80:
            return "Bradicardia"
        elif 80 <= bpm <= 130:
            return "Normal"
        else:
            return "Taquicardia"

    # Niños de 4 a 5 años (Ambos sexos)
    elif 4 <= age <= 5:
        if bpm < 80:
            return "Bradicardia"
        elif 80 <= bpm <= 120:
            return "Normal"
        else:
            return "Taquicardia"

    # Niños de 6 a 12 años (Ambos sexos)
    elif 6 <= age <= 12:
        if bpm < 70:
            return "Bradicardia"
        elif 70 <= bpm <= 110:
            return "Normal"
        else:
            return "Taquicardia"

    # Pacientes de 13 años en adelante (Adolescentes y Adultos)
    # Aquí sí hay diferencia si es Hombre o Mujer.
    else:
        # Si el paciente es Hombre (Aplica para 13 años o más)
        if sex == "Hombre":
            if bpm < 60:
                return "Bradicardia"
            elif 60 <= bpm <= 100:
                return "Normal"
            else:
                return "Taquicardia"

        # Si el paciente es Mujer (Aplica para 13 años o más, no embarazada)
        else:
            if bpm < 60:
                return "Bradicardia"
            elif 62 <= bpm <= 102:
                return "Normal"
            else:
                return "Taquicardia"


# MANEJO DE INTERFAZ DE USUARIO
def clear_screen():
    """Wipes the window clean before drawing a new screen."""
    for widget in root.winfo_children():
        widget.destroy()


def show_registration_screen():
    """
    SCREEN 1: Draws the initial registration form for inputting patient information.
    """
    clear_screen()

    # Título de la ventana
    ctk.CTkLabel(root, text="Registro de Paciente", font=("Arial", 22, "bold")).pack(
        pady=20
    )

    # Contenedor
    frame = ctk.CTkFrame(root)
    frame.pack(pady=10, padx=30, fill="both", expand=True)

    # Ingreso de nombre
    ctk.CTkLabel(frame, text="Nombre Completo:").pack(pady=5)
    entry_name = ctk.CTkEntry(frame, width=250)
    entry_name.pack(pady=5)

    # Ingreso de edad
    ctk.CTkLabel(frame, text="Edad:").pack(pady=5)
    entry_age = ctk.CTkEntry(frame, width=250)
    entry_age.pack(pady=5)

    # Ingreso de sexo
    ctk.CTkLabel(frame, text="Sexo:").pack(pady=5)
    dropdown_sex = ctk.CTkComboBox(frame, values=["Masculino", "Femenino"], width=250)
    dropdown_sex.pack(pady=5)

    # Checkbox de embarazo
    ctk.CTkLabel(frame, text="Si está embarazada, seleccione la etapa:").pack(pady=2)
    dropdown_pregnant = ctk.CTkComboBox(
        frame,
        values=[
            "No embarazada",
            "Primer trimestre",
            "Segundo trimestre",
            "Tercer trimestre",
            "Embarazo avanzado",
        ],
        width=250,
    )
    dropdown_pregnant.pack(pady=5)

    def toggle_pregnancy(choice):
        """
        A watcher function. If the user picks 'Femenino', the pregnancy checkbox
        appears on screen. If they choose 'Masculino', it instantly hides itself.
        """
        if choice == "Femenino":
            dropdown_pregnant.pack(pady=5)
        else:
            dropdown_pregnant.pack_forget()
            dropdown_pregnant.set("No embarazada")

    # Conectar el menú dropdown a la función de arriba
    dropdown_sex.configure(command=toggle_pregnancy)

    def handle_continue():
        """
        Runs when the user clicks 'Continuar'. Reads the entries, saves them to our
        global variables, verifies fields aren't blank, and triggers Screen 2.
        """
        global user_name, user_age, user_sex, user_pregnant

        # .get() lee lo que está dentro de la caja de texto, .strip() elimina espacios en blanco
        user_name = entry_name.get().strip()
        user_age = entry_age.get().strip()
        user_sex = dropdown_sex.get()
        user_pregnant = dropdown_pregnant.get()

        # Avanzar a la siguiente pantalla sólo si los campos han sido llenados debidamente
        if user_name and user_age and user_sex:
            show_dashboard_screen()

    ctk.CTkButton(root, text="Continuar", command=handle_continue).pack(pady=20)


def show_dashboard_screen():
    """
    SCREEN 2: Displays the specific historical data list for the patient
    and holds the delete/measurement controls.
    """
    clear_screen()

    # Título
    ctk.CTkLabel(
        root, text=f"Historial de {user_name}", font=("Arial", 20, "bold")
    ).pack(pady=15)

    # Tabla vacia para mostrar los registros
    listbox = Listbox(
        root,
        font=("Courier New", 12),
        bg="#222222",
        fg="white",
        selectbackground="#2ecc71",
    )
    listbox.pack(padx=20, pady=10, fill="both", expand=True)

    excel_rows = {}  # Un diccionario en blanco para llevar un control sobre los índices de la tabla vacia y los indices de la tabla de excel

    def load_history():
        """
        Opens the Excel file, loops through every single row, filters rows that
        match the patient's name, and inserts them dynamically into the visual list box.
        """
        listbox.delete(0, END)  # Limpiar la tabla de la aplicación antes de recargar
        excel_rows.clear()  # Reiniciar el indice de la tabla de Excel

        wb = load_workbook(xlfile)
        ws = wb.active

        current_item_index = 0
        # Para listar los registros del Excel en la aplicación, hay que iterar por ellos
        # Se comienza desde la fila 2 ya que la fila 1 solo tiene los títulos de las columnas
        for row in range(2, ws.max_row + 1):
            # Se convierte a mayúscula el nombre para evitar inconsistencias entre registros
            if str(ws.cell(row=row, column=1).value).upper() == user_name.upper():
                bpm = ws.cell(row=row, column=5).value
                risk = ws.cell(row=row, column=6).value
                # Formatear el texto para insertarlo en la tabla de la aplicación. En el Excel, se agrega la información de forma normal
                listbox.insert(END, f"  BPM: {bpm} | Riesgo: {risk}")
                excel_rows[current_item_index] = row
                current_item_index += 1

        # Si no hay registros de ese usuario, se muestra una leyenda
        if listbox.size() == 0:
            listbox.insert(END, " No hay registros para este paciente.")

    # Se llama a la función cada que se cambia a esta pantalla
    load_history()

    def delete_record():
        """
        Runs when the 'Eliminar Registro' button is clicked. Looks up which line
        you highlighted, matches it to Excel, deletes it from disk, and reloads.
        """
        selected = listbox.curselection()  # Guardar el registro seleccionado
        if selected and selected[0] in excel_rows:
            wb = load_workbook(xlfile)
            ws = wb.active

            # Ubicar la fila en Excel y borrarla
            ws.delete_rows(excel_rows[selected[0]], 1)
            wb.save(xlfile)

            # Recargar el historial una vez borrado el registro
            load_history()

    # Layout de los botones
    btn_frame = ctk.CTkFrame(root, fg_color="transparent")
    btn_frame.pack(pady=10)

    # Botón de eliminar registro
    ctk.CTkButton(
        btn_frame, text="Eliminar Registro", fg_color="#e74c3c", command=delete_record
    ).pack(side="left", padx=10)

    # Botón para comenzar una nueva medición
    ctk.CTkButton(
        btn_frame,
        text="Nueva Medición",
        fg_color="#2ecc71",
        command=lambda: open_popup(load_history),
    ).pack(side="left", padx=10)

    # Botón para regresar a la pantalla anterior (Registro)
    ctk.CTkButton(
        root, text="Regresar", fg_color="gray", command=show_registration_screen
    ).pack(pady=10)


# POP-UP PARA HARDWARE
def open_popup(refresh_callback):
    """
    Creates a distinct temporary popup overlay window to ask if the
    sensor machine is turned on before starting the Bluetooth antenna.
    """
    popup = ctk.CTkToplevel(root)
    popup.title("Hardware")
    popup.geometry("380x200")

    # Esperar a que el pop-up aparezca y cambiar el focus a este
    popup.wait_visibility()
    popup.grab_set()

    ctk.CTkLabel(
        popup, text="¿El dispositivo está encendido?", font=("Arial", 15, "bold")
    ).pack(pady=15)
    lbl_status = ctk.CTkLabel(
        popup, text="", font=("Arial", 13, "italic"), text_color="gray"
    )
    lbl_status.pack(pady=5)

    def press_no():
        lbl_status.configure(
            text="Por favor, encienda el dispositivo.", text_color="#e74c3c"
        )

    def press_yes():
        # Dar un feedback inmediato para hacerle saber al usuario que la aplicación no ha crasheado
        lbl_status.configure(
            text="Conectando al sensor... Espere.", text_color="#3498db"
        )
        root.update()

        # Se llama a la función que conecta con el módulo bluetooth
        bpm_result = read_bluetooth_sensor()

        if bpm_result == "FAIL":
            # Hacerle saber al usuario que el dispositivo esta desconectado, fuera de rango o apagado
            lbl_status.configure(
                text="Dispositivo apagado o fuera de rango.", text_color="#e74c3c"
            )
        else:
            # Se captura la medición
            lbl_status.configure(text="¡Medición exitosa!", text_color="#2ecc71")
            root.update()

            # Calcular el riesgo en base a la medición recibida
            risk_status = calculate_risk(bpm_result, user_age, user_sex, user_pregnant)
            wb = load_workbook(xlfile)
            ws = wb.active
            ws.append(
                [
                    user_name,
                    int(user_age),
                    user_sex,
                    user_pregnant,
                    int(bpm_result),
                    risk_status,
                ]
            )
            wb.save(xlfile)

            # Mostrar una leyenda que la medición fue exitosa
            root.after(1200, lambda: [popup.destroy(), refresh_callback()])

    btn_box = ctk.CTkFrame(popup, fg_color="transparent")
    btn_box.pack(pady=15)

    # Botón para realizar la medición
    ctk.CTkButton(
        btn_box, text="Sí", fg_color="#2ecc71", width=90, command=press_yes
    ).pack(side="left", padx=10)

    # Botón para hacerle saber al usuario de conectar el dispositivo
    ctk.CTkButton(
        btn_box, text="No", fg_color="#e74c3c", width=90, command=press_no
    ).pack(side="left", padx=10)


# CONEXIÓN BLUETOOTH
def read_bluetooth_sensor():
    """
    The hardware workhorse. It reaches out, connects to THE HM-10 address,
    listens for a line containing 'FC=', strips out the numeric digits,
    disconnects, and returns that value.
    """
    final_bpm = "FAIL"  # Respuesta predeterminada en caso de error

    async def fetch_data():
        """An isolated helper script block capable of handling hardware notifications."""
        nonlocal final_bpm

        def packet_inspector(sender, data):
            """An internal parser. It intercepts the raw binary text streaming from your HM-10."""
            nonlocal final_bpm
            # Decodificar información binaria a texto legible
            raw_text = data.decode("utf-8").strip().upper()
            # Se extrae la medición mediante uso de expresiones regulares (regex)
            match = re.search(r"FC=(\d+)", raw_text)
            if match:
                final_bpm = match.group(1)  # Se toman solo los números

        try:
            # Conectar con el módulo físico con un límite de 15 segundos para lograrlo
            async with BleakClient(DEVICE_ADDRESS, timeout=15.0) as client:
                if client.is_connected:
                    # El programa se suscribe al canal del módulo de donde se transmiten los datos
                    await client.start_notify(HM10_UUID, packet_inspector)
                    # Una vez conectado, espera un máximo de 30 segundos para recibir la medición
                    for _ in range(30):
                        if final_bpm != "FAIL":
                            break  # Se recibe la medida y se rompe el ciclo
                        await asyncio.sleep(
                            1
                        )  # Se espera un segundo y vuelve a revisar si hay nueva información
        except Exception:
            final_bpm = "FAIL"

    # El proceso se tiene que manejar de manera asincrona para que el programa entero no se congele mientras espera
    # una respuesta del módulo bluetooth
    asyncio.run(fetch_data())
    return final_bpm


if __name__ == "__main__":
    # Se crea la pantalla
    root = ctk.CTk()
    root.title("Monitor BPM")
    root.geometry("650x500")

    # El programa muestra la pantalla de registro al inicio
    show_registration_screen()

    # Mantiene la ventana (y el programa) abiertos hasta que esta se cierre o se interrumpa el programa
    root.mainloop()

# ----IMPORTANTÍSIMO Y BORRAR ESTO CUANDO LO ENTREGUEN------
#
# Hay dos conceptos importantes de este código que deben saber para comprender un poco mejor que esta haciendo el programa
# 1. Regex o Expresiones regulares
# 2. Funcionamiento asíncrono
#
# Son sencillos de entender.

# Una expresión regular es básicamente un patrón que quieren buscar en un palabra. Si ven se van a la línea 472 veran que:
#   match = re.search(r"FC=(\d+)", raw_text)
# Sabemos que el módulo BT manda tres cosas cuando se enciende: LISTO, LEYENDO y FC=<número>
# LISTO y LEYENDO no nos importa mucho porque simplemente tomamos ese output y los mostramos en texto a la aplicación
# La parte importante es FC=<número> pero como verán, ese número es texto y esta embedido en otro texto entonces tenemos que sacar el número de ahí
#
# En la línea mencionada antes le estamos diciendo a Python:
# "Oye, de ese mensaje en texto plano que te llegó (raw_text), ubica el patrón 'FC=' y traete cualquier número que encuentres
# después de ese patrón, con todos los digitos que tenga(/d+)"
# Eso es una expresión regular, buscar un patrón (números, dígitos, símbolos) en otra expresión o palabra
#
#
# Lo siguiente ya no tan importan son las funciones asíncronas.
# Si tienen poca experiencia en Python recordarán que el código que escriben se ejecuta de manera consecutiva.
# Es decir, la línea 1 se ejecuta primero, después la línea 2, después la línea 3 y así ad infinitum
#
# Entonces, en base a ese proceso podemos decir que las líneas de código o tareas tienen que esperar a que su linea anterior termine de ejecutarse
# antes empezar a ejecutarse a sí mismas. Esto significa que la línea 3 no se puede ejecutar hasta que las líneas 1 y 2 se ejecuten. Eso es funcionalidad síncrona,
# que las tareas del código se ejecutan de forma sincronizada, una tras de otra.

# Ahora imaginen que están en su navegador web conectados al Internet del Tec y quieren buscan una página web pero es un día común y hay conexión lenta.
# Ustedes ven que su navegador se queda cargando la página pero pueden cambiar de pestaña, pueden cambiar las opciones de su navegador,
# pueden cerrar el navegador, pueden seguir operando el navegador a pesar de que su solicitud siga cargando o que de plano falle, eso es funcionalidad asíncrona.
# Asíncrono es cuando tenemos varias tareas que se ejecutan independientemente de la otra. Si una tarea falla, no tenemos que preocuparnos de que nuestro sistema
# entero falle en su totalidad por ello.

# Si ven la línea 462 dentro de la función para la conexión bluetooth, verán que la función de fetch_data() comienza con un "async".
# Eso significa que si no recibimos nada de información, el programa no se quedará congelado por siempre esperando sino
# que solo nos enviará un error pero nuestro programa principal seguira corriendo.

# Vayamos ahora a la línea 478, tenemos otro async. Por como funciona su módulo, nos tenemos que suscribir a un canal del módulo que se encarga
# se enviar la información de las mediciones. Si vieron hay dos variables en mayúsculas al principio del archivo: DEVICE_ADDRESS y HM10_UUID. DEVICE_ADDRESS
# DEVICE_ADDRESS es la dirección MAC de su dispositivo, la dirección MAC es un identificador único que cada dispositivo electrónico en el mundo tiene. HM10_UUID es
# el identificador del canal de su módulo que envía la información. Con esos dos datos se realiza la conexión BT (cualquier conexión BT btw)
# Regresemos al async, aquí es la misma lógica, tenemos que crear una función asincrona para que en caso de la conexión con el módulo falle
# por cualquier cosa, el programa no se quedé congelado.
# Si ven la línea 481 verán que hay una palabra "await", eso significa que le estamos diciendo a nuestra función asíncrona que se espere hasta que reciba una
# medición para procesarla
